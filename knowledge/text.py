"""Helpers for extracting text from office documents."""

from __future__ import annotations

import io
import os
import re
import subprocess
import tempfile
from pathlib import Path
from typing import Iterable

import structlog
from pypdf import PdfReader


logger = structlog.get_logger(__name__)

TEXT_LIKE_MIME_TYPES = {
    "application/json",
    "application/xml",
    "text/csv",
}

PDF_MIME_TYPES = {"application/pdf"}

DOCX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-word.document.macroenabled.12",
}

DOC_MIME_TYPES = {
    "application/msword",
    "application/ms-word",
    "application/vnd.ms-word",
    "application/vnd.ms-word.document.macroenabled.12",
}

XLSX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroenabled.12",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.template",
    "application/vnd.ms-excel.sheet.binary.macroenabled.12",
}

XLS_MIME_TYPES = {
    "application/vnd.ms-excel",
    "application/ms-excel",
    "application/xls",
    "application/vnd.ms-excel.sheet.macroenabled.12",
}

try:  # pragma: no cover - optional dependency in some environments
    import docx2txt  # type: ignore
except Exception:  # noqa: BLE001
    docx2txt = None

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook  # type: ignore
except Exception:  # noqa: BLE001
    load_workbook = None

try:  # pragma: no cover - optional (only needed for legacy XLS files)
    import xlrd  # type: ignore
except Exception:  # noqa: BLE001
    xlrd = None


def extract_docx_text(data: bytes) -> str:
    """Return plain text extracted from a DOCX payload."""

    if docx2txt is None:
        logger.warning("docx_extract_unavailable", reason="docx2txt_not_installed")
        return ""

    with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
        tmp.write(data)
        tmp.flush()
        tmp_path = Path(tmp.name)

    try:
        text = docx2txt.process(tmp_path)
    except Exception as exc:  # noqa: BLE001
        logger.warning("docx_extract_failed", error=str(exc))
        text = ""
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:  # pragma: no cover - cleanup best effort
            pass

    return (text or "").strip()


def _run_textract(path: Path) -> str:
    try:
        import textract  # type: ignore
    except Exception:  # pragma: no cover - optional dependency
        return ""

    try:
        blob = textract.process(str(path))
    except Exception as exc:  # noqa: BLE001
        logger.debug("doc_extract_textract_failed", error=str(exc))
        return ""
    return blob.decode("utf-8", errors="ignore").strip()


def _run_antiword(path: Path) -> str:
    try:
        result = subprocess.run(
            ["antiword", str(path)],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
    except FileNotFoundError:  # pragma: no cover - optional binary
        logger.debug("doc_extract_antiword_missing")
        return ""
    except subprocess.CalledProcessError as exc:  # noqa: BLE001
        logger.debug("doc_extract_antiword_failed", error=str(exc))
        return ""
    return result.stdout.decode("utf-8", errors="ignore").strip()


def extract_doc_text(data: bytes) -> str:
    """Return plain text extracted from a legacy DOC payload."""

    with tempfile.NamedTemporaryFile(delete=False, suffix=".doc") as tmp:
        tmp.write(data)
        tmp.flush()
        tmp_path = Path(tmp.name)

    try:
        text = _run_textract(tmp_path)
        if text:
            return text

        text = _run_antiword(tmp_path)
        if text:
            return text

        raw = tmp_path.read_bytes()
        decoded = raw.decode("utf-8", errors="ignore")
        cleaned = decoded.replace("\x00", " ")
        cleaned = re.sub(r"\s{2,}", " ", cleaned)
        return cleaned.strip()
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:  # pragma: no cover - cleanup best effort
            pass


def _clean_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return ("%s" % value).strip()
    return str(value).strip()


def _render_sheet_rows(rows: Iterable[Iterable[object]]) -> list[str]:
    rendered: list[str] = []
    for row in rows:
        cleaned_cells: list[str] = []
        for cell in row:
            cleaned = _clean_cell_value(cell)
            if cleaned:
                cleaned_cells.append(cleaned)
        if not cleaned_cells:
            continue
        rendered.append(" \u2014 ".join(cleaned_cells))
    return rendered


def _extract_with_textract(data: bytes, suffix: str) -> str:
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(data)
        tmp.flush()
        tmp_path = Path(tmp.name)
    try:
        text = _run_textract(tmp_path)
        return text.strip()
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:  # pragma: no cover
            pass


def extract_xlsx_text(data: bytes) -> str:
    """Return textual content rendered from an XLSX/XLSM workbook."""

    if load_workbook is None:
        logger.warning("xlsx_extract_unavailable", reason="openpyxl_missing")
        return _extract_with_textract(data, ".xlsx")

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("xlsx_extract_open_failed", error=str(exc))
        return _extract_with_textract(data, ".xlsx")

    try:
        sheet_chunks: list[str] = []
        for worksheet in workbook.worksheets:
            sheet_lines = _render_sheet_rows(worksheet.iter_rows(values_only=True))
            if not sheet_lines:
                continue
            title = (worksheet.title or "Sheet").strip()
            sheet_chunks.append(f"[{title}]")
            sheet_chunks.extend(sheet_lines)
        rendered = "\n".join(chunk.strip() for chunk in sheet_chunks if chunk.strip()).strip()
        if rendered:
            return rendered
    finally:
        try:
            workbook.close()
        except Exception:  # pragma: no cover - best effort
            pass

    return _extract_with_textract(data, ".xlsx")


def extract_xls_text(data: bytes) -> str:
    """Return textual content from a legacy XLS workbook."""

    if xlrd is not None:
        try:
            book = xlrd.open_workbook(file_contents=data)
        except Exception as exc:  # noqa: BLE001
            logger.warning("xls_extract_open_failed", error=str(exc))
        else:
            try:
                sheet_chunks: list[str] = []
                for sheet in book.sheets():
                    rows: list[list[object]] = []
                    for row_index in range(sheet.nrows):
                        rows.append(sheet.row_values(row_index))
                    sheet_lines = _render_sheet_rows(rows)
                    if not sheet_lines:
                        continue
                    title = (sheet.name or "Sheet").strip()
                    sheet_chunks.append(f"[{title}]")
                    sheet_chunks.extend(sheet_lines)
                rendered = "\n".join(chunk.strip() for chunk in sheet_chunks if chunk.strip()).strip()
                if rendered:
                    return rendered
            finally:
                try:
                    book.release_resources()
                except Exception:  # pragma: no cover
                    pass

    return _extract_with_textract(data, ".xls")


def extract_pdf_text(data: bytes) -> str:
    """Return plain text extracted from a PDF payload using :mod:`pypdf`."""

    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception as exc:  # noqa: BLE001
        logger.debug("pdf_extract_open_failed", error=str(exc))
        return ""

    chunks: list[str] = []
    for index, page in enumerate(reader.pages):
        try:
            extracted = page.extract_text() or ""
        except Exception as exc:  # noqa: BLE001
            logger.debug("pdf_extract_page_failed", page=index, error=str(exc))
            continue
        if extracted:
            chunks.append(extracted.strip())

    if chunks:
        return "\n\n".join(part for part in chunks if part).strip()

    return _extract_with_textract(data, ".pdf")


def _decode_text_payload(payload: bytes) -> str:
    if not payload:
        return ""
    text = payload.decode("utf-8", errors="ignore").replace("\x00", " ").strip()
    if text:
        return text
    fallback = payload.decode("latin-1", errors="ignore").replace("\x00", " ").strip()
    return fallback


def extract_best_effort_text(name: str, content_type: str | None, payload: bytes) -> str:
    """Return textual content extracted from an arbitrary binary blob.

    Falls back to heuristics when the payload does not match a known office format.
    """

    lowered = (content_type or "").lower()
    safe_name = (name or "").lower()
    if lowered.startswith("text/") or lowered in TEXT_LIKE_MIME_TYPES:
        return _decode_text_payload(payload)
    if lowered in PDF_MIME_TYPES or safe_name.endswith(".pdf"):
        return extract_pdf_text(payload)
    if lowered in DOCX_MIME_TYPES or safe_name.endswith(".docx"):
        return extract_docx_text(payload)
    if lowered in DOC_MIME_TYPES or safe_name.endswith(".doc"):
        return extract_doc_text(payload)
    if lowered in XLSX_MIME_TYPES or safe_name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xlsb")):
        return extract_xlsx_text(payload)
    if lowered in XLS_MIME_TYPES or safe_name.endswith((".xls", ".xlt", ".xlm", ".xla", ".xlw")):
        return extract_xls_text(payload)

    decoded = _decode_text_payload(payload)
    # Heuristic: treat as text only if it contains a reasonable number of printable chars
    printable_ratio = 0.0
    if decoded:
        total = len(decoded)
        if total:
            printable = sum(1 for ch in decoded if ch.isprintable() or ch.isspace())
            printable_ratio = printable / total
    if decoded and printable_ratio >= 0.7:
        cleaned = re.sub(r"\s{2,}", " ", decoded)
        return cleaned.strip()
    return ""

"""Knowledge service API endpoints."""

from typing import Any, Literal
import asyncio
from uuid import uuid4
from contextlib import suppress

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Form, Response
from fastapi.responses import ORJSONResponse

from backend.auth import require_admin, require_super_admin, resolve_admin_project
from backend.utils.project import normalize_project
from backend.api.utils import build_download_url
from knowledge.tasks import queue_auto_description
from settings import MongoSettings
from observability.logging import get_logger
from backend.knowledge.schemas import (
    IntelligentProcessingPromptPayload,
    KnowledgeCreate,
    KnowledgeDeduplicate,
    KnowledgePriorityPayload,
    KnowledgeQAPayload,
    KnowledgeQAReorderPayload,
    KnowledgeServiceConfig,
    KnowledgeServiceRunRequest,
    KnowledgeUnansweredClearPayload,
    KnowledgeUpdate,
)
from knowledge_service.configuration import (
    ALLOWED_MODES as KNOWLEDGE_SERVICE_ALLOWED_MODES,
    DEFAULT_MODE as KNOWLEDGE_SERVICE_DEFAULT_MODE,
    DEFAULT_PROCESSING_PROMPT as KNOWLEDGE_SERVICE_DEFAULT_PROMPT,
    KEY as KNOWLEDGE_SERVICE_KEY,
    MANUAL_MODE_MESSAGE as KNOWLEDGE_SERVICE_MANUAL_MESSAGE,
)

from knowledge_service.runner import KnowledgeServiceRunner

logger = get_logger(__name__)

_KNOWN_KNOWLEDGE_SOURCES = {
    "uploaded_files",
    "website_content",
    "bitrix_knowledge_base",
    "manual_input",
}
_DEFAULT_KNOWLEDGE_PRIORITY = [
    "manual_input",
    "uploaded_files",
    "website_content",
    "bitrix_knowledge_base",
]

# Helper for QA upload
async def _read_qa_upload(file: UploadFile) -> list[dict[str, Any]]:
    """Parse uploaded file (CSV, Excel, or text) into QA pairs."""
    import csv
    import io
    import html
    import unicodedata
    
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File is empty")
        
    pairs = []
    
    # 1. Excel
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            from openpyxl import load_workbook
            buffer = io.BytesIO(content)
            wb = load_workbook(buffer, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            
            # Skip header if present
            header = rows[0]
            start_idx = 0
            if header and isinstance(header[0], str) and (
                "question" in header[0].lower() or "вопрос" in header[0].lower()
            ):
                start_idx = 1
                
            for row in rows[start_idx:]:
                if len(row) < 2:
                    continue
                q, a = row[0], row[1]
                p = row[2] if len(row) > 2 else 0
                if q and a:
                    pairs.append({"question": str(q), "answer": str(a), "priority": int(p) if p else 0})
        except ImportError:
            logger.warning("openpyxl_missing_for_excel_import")
            raise HTTPException(status_code=400, detail="Excel support not available")
        except Exception as exc:
            logger.error("excel_parse_failed", error=str(exc))
            raise HTTPException(status_code=400, detail="Failed to parse Excel file")

    # 2. CSV
    elif filename.endswith(".csv") or b"," in content or b";" in content or b"\t" in content:
        try:
            text = content.decode("utf-8-sig")
            # Try to sniff dialect
            try:
                dialect = csv.Sniffer().sniff(text[:1024], delimiters=";, \t")
            except csv.Error:
                dialect = None
                
            f = io.StringIO(text)
            reader = csv.reader(f, dialect) if dialect else csv.reader(f)
            
            rows = list(reader)
            if not rows:
                return []
                
            start_idx = 0
            if rows and (
                "question" in rows[0][0].lower() or "вопрос" in rows[0][0].lower()
            ):
                start_idx = 1
                
            for row in rows[start_idx:]:
                if len(row) < 2:
                    continue
                q, a = row[0], row[1]
                p = row[2] if len(row) > 2 else 0
                if q and a:
                    pairs.append({"question": q, "answer": a, "priority": int(p) if p else 0})
        except Exception as exc:
            # Fallback to text parser if CSV fails
            logger.warning("csv_parse_failed", error=str(exc))
            pass

    # 3. Text / Fallback
    if not pairs:
        try:
            text = content.decode("utf-8", errors="ignore")
        except Exception:
            text = ""
            
        current_q = None
        current_a = None
        
        for line in text.splitlines():
            line = line.strip()
            if not line:
                if current_q and current_a:
                    pairs.append({"question": current_q, "answer": current_a, "priority": 0})
                    current_q = None
                    current_a = None
                continue
                
            if line.lower().startswith("q:") or line.lower().startswith("в:"):
                if current_q and current_a:
                    pairs.append({"question": current_q, "answer": current_a, "priority": 0})
                    current_a = None
                current_q = line.split(":", 1)[1].strip()
            elif line.lower().startswith("a:") or line.lower().startswith("o:") or line.lower().startswith("о:"):
                current_a = line.split(":", 1)[1].strip()
                
        if current_q and current_a:
            pairs.append({"question": current_q, "answer": current_a, "priority": 0})

    # Cleanup and validation
    cleaned_pairs = []
    for p in pairs:
        q = str(p["question"]).strip()
        a = str(p["answer"]).strip()
        
        # Truncate
        if len(q) > 1000:
            q = q[:1000]
        if len(a) > 10000:
            a = a[:10000]
            
        # HTML escape
        q = html.escape(q)
        
        # Unicode normalization
        q = unicodedata.normalize("NFKC", q)
        a = unicodedata.normalize("NFKC", a)
        
        if q and a:
            cleaned_pairs.append({
                "question": q,
                "answer": a,
                "priority": p.get("priority", 0)
            })
            
    return cleaned_pairs

router = APIRouter(prefix="/knowledge", tags=["knowledge"])


@router.get("", response_class=ORJSONResponse)
async def knowledge_list(
    request: Request,
    project: str | None = None,
    domain: str | None = None,
    limit: int = 100,
    offset: int = 0,
    search: str | None = None,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """List knowledge items for a project."""
    project_scope = resolve_admin_project(request, project)
    if not project_scope and not domain:
        # Fallback to primary project if available
        identity = request.state.admin
        if identity and identity.primary_project:
            project_scope = identity.primary_project

    items = await request.state.mongo.list_knowledge(
        project=project_scope,
        domain=domain,
        limit=limit,
        offset=offset,
        search=search,
    )
    return ORJSONResponse({"items": [item.model_dump(by_alias=True) for item in items]})


@router.post("", response_class=ORJSONResponse)
async def knowledge_create(
    request: Request,
    payload: KnowledgeCreate,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Create a new knowledge item."""
    project_scope = resolve_admin_project(request, payload.project)
    item = await request.state.mongo.create_knowledge(
        project=project_scope,
        content=payload.content,
        name=payload.name,
        domain=payload.domain,
        description=payload.description,
        url=payload.url,
    )
    return ORJSONResponse(item.model_dump(by_alias=True))


@router.post("/upload", response_class=ORJSONResponse, status_code=201)
async def knowledge_upload(
    request: Request,
    project: str = Form(...),
    description: str | None = Form(None),
    name: str | None = Form(None),
    url: str | None = Form(None),
    file: UploadFile = File(...),
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Upload a binary document to GridFS and store metadata."""

    project_name = resolve_admin_project(request, project, required=True)
    # Get project model for domain
    project_model = await request.state.mongo.get_project(project_name)
    
    mongo_cfg = MongoSettings()
    collection = getattr(request.state, "documents_collection", mongo_cfg.documents)

    filename = (name or file.filename or "").strip()
    if not filename:
        raise HTTPException(status_code=400, detail="File name is required")

    payload = await file.read()
    if not payload:
        raise HTTPException(status_code=400, detail="File is empty")

    description_input = (description or "").strip()
    auto_description_pending = False
    status_message = None
    if description_input:
        description_value = description_input
    else:
        description_value = ""
        auto_description_pending = True
        status_message = "Автоописание в очереди"

    try:
        file_id = await request.state.mongo.upload_document(
            file_name=filename,
            file=payload,
            documents_collection=collection,
            description=description_value,
            url=url,
            content_type=file.content_type,
            project=project_name,
            domain=project_model.domain if project_model else None,
        )
        download_url = build_download_url(request, file_id)
        
        # Update with download URL
        await request.state.mongo.db[collection].update_one(
            {"fileId": file_id},
            {"$set": {"url": download_url, "content_type": file.content_type}},
            upsert=False,
        )
        
        # Update auto-description status
        await request.state.mongo.db[collection].update_one(
            {"fileId": file_id},
            {"$set": {"autoDescriptionPending": auto_description_pending}},
            upsert=False,
        )
        
        if auto_description_pending:
            await request.state.mongo.update_document_status(
                collection,
                file_id,
                "pending_auto_description",
                status_message,
            )
            queue_auto_description(file_id, project_name)
        else:
            await request.state.mongo.update_document_status(
                collection,
                file_id,
                "ready",
                "Описание задано вручную",
            )
            
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return ORJSONResponse(
        {
            "file_id": file_id,
            "name": filename,
            "project": project_name,
            "download_url": download_url,
            "content_type": file.content_type,
            "description": description_value,
            "auto_description_pending": auto_description_pending,
            "status": "pending_auto_description" if auto_description_pending else "ready",
            "status_message": status_message,
        }
    )


@router.post("/reindex", response_class=ORJSONResponse)
async def knowledge_reindex(
    request: Request,
    _: Any = Depends(require_super_admin),
) -> ORJSONResponse:
    """Trigger reindexing of vector store."""
    loop = asyncio.get_running_loop()

    def _run_update() -> None:
        from worker import update_vector_store
        update_vector_store()

    loop.run_in_executor(None, _run_update)
    return ORJSONResponse({"status": "queued"})


@router.delete("", response_class=ORJSONResponse)
async def knowledge_clear(
    request: Request,
    project: str | None = None,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Remove documents from the knowledge base (optionally scoped to a project)."""
    project_scope = resolve_admin_project(request, project)
    
    mongo_cfg = MongoSettings()
    collection = getattr(request.state, "documents_collection", mongo_cfg.documents)
    
    summary = await request.state.mongo.delete_documents(collection, project_scope)

    loop = asyncio.get_running_loop()

    def _refresh() -> None:
        from worker import update_vector_store
        update_vector_store()

    loop.run_in_executor(None, _refresh)
    return ORJSONResponse({"project": project_scope, **summary})


@router.get("/priority", response_class=ORJSONResponse)
async def knowledge_priority_get(
    request: Request,
    project: str | None = None,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Get knowledge source priority."""
    project_scope = resolve_admin_project(request, project)
    try:
        stored_order = await request.state.mongo.get_knowledge_priority(project_scope)
    except Exception as exc:
        logger.error("knowledge_priority_fetch_failed", project=project_scope, error=str(exc))
        raise HTTPException(status_code=500, detail="Failed to load knowledge priority") from exc

    available_sources = list(_KNOWN_KNOWLEDGE_SOURCES)
    normalized = []
    for entry in stored_order or []:
        candidate = str(entry).strip()
        if candidate in available_sources and candidate not in normalized:
            normalized.append(candidate)
    if not normalized:
        normalized = [source for source in _DEFAULT_KNOWLEDGE_PRIORITY if source in available_sources]
    if not normalized:
        normalized = available_sources

    return ORJSONResponse(
        {
            "order": normalized,
            "available": available_sources,
            "project": project_scope,
        }
    )


@router.post("/priority", response_class=ORJSONResponse)
async def knowledge_priority_set(
    request: Request,
    payload: KnowledgePriorityPayload,
    project: str | None = None,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Set knowledge source priority."""
    project_scope = resolve_admin_project(request, project)

    available_sources = list(_KNOWN_KNOWLEDGE_SOURCES)
    lookup = {source.lower(): source for source in available_sources}
    normalized: list[str] = []
    for entry in payload.order or []:
        candidate = lookup.get(str(entry or "").strip().lower())
        if candidate and candidate not in normalized:
            normalized.append(candidate)

    for fallback in _DEFAULT_KNOWLEDGE_PRIORITY:
        if fallback in available_sources and fallback not in normalized:
            normalized.append(fallback)

    if not normalized:
        normalized = available_sources

    try:
        await request.state.mongo.set_knowledge_priority(project_scope, normalized)
    except Exception as exc:
        logger.error("knowledge_priority_save_failed", project=project_scope, error=str(exc))
        raise HTTPException(status_code=500, detail="Failed to save knowledge priority") from exc

    return ORJSONResponse({"order": normalized, "project": project_scope})
@router.get("/{item_id}", response_class=ORJSONResponse)
async def knowledge_get(
    request: Request,
    item_id: str,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Get a knowledge item by ID."""
    item = await request.state.mongo.get_knowledge(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Knowledge item not found")
    
    # Check access
    resolve_admin_project(request, item.project)
    
    return ORJSONResponse(item.model_dump(by_alias=True))


@router.patch("/{item_id}", response_class=ORJSONResponse)
async def knowledge_update(
    request: Request,
    item_id: str,
    payload: KnowledgeUpdate,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Update a knowledge item."""
    # Check existence and access first
    existing = await request.state.mongo.get_knowledge(item_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Knowledge item not found")
    
    resolve_admin_project(request, existing.project)
    
    update_data = payload.model_dump(exclude_unset=True)
    if not update_data:
        return ORJSONResponse(existing.model_dump(by_alias=True))

    item = await request.state.mongo.update_knowledge(item_id, update_data)
    return ORJSONResponse(item.model_dump(by_alias=True))


@router.delete("/{item_id}", response_class=ORJSONResponse)
async def knowledge_delete(
    request: Request,
    item_id: str,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Delete a knowledge item."""
    existing = await request.state.mongo.get_knowledge(item_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Knowledge item not found")
        
    resolve_admin_project(request, existing.project)
    
    await request.state.mongo.delete_knowledge(item_id)
    return ORJSONResponse({"status": "deleted"})


@router.post("/deduplicate", response_class=ORJSONResponse)
async def knowledge_deduplicate(
    request: Request,
    payload: KnowledgeDeduplicate,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Find duplicate knowledge items."""
    project_scope = resolve_admin_project(request, payload.project)
    duplicates = await request.state.mongo.find_duplicate_knowledge(project_scope)
    return ORJSONResponse({"duplicates": duplicates})


# QA Endpoints

@router.get("/qa", response_class=ORJSONResponse)
async def knowledge_qa_list(
    request: Request,
    project: str | None = None,
    limit: int = 100,
    offset: int = 0,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """List QA pairs."""
    project_scope = resolve_admin_project(request, project)
    items = await request.state.mongo.list_qa_pairs(
        project=project_scope,
        limit=limit,
        offset=offset,
    )
    return ORJSONResponse({"items": [item.model_dump(by_alias=True) for item in items]})


@router.post("/qa", response_class=ORJSONResponse)
async def knowledge_qa_create(
    request: Request,
    payload: KnowledgeQAPayload,
    project: str | None = None,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Create a QA pair."""
    project_scope = resolve_admin_project(request, project, required=True)
    item = await request.state.mongo.create_qa_pair(
        project=project_scope,
        question=payload.question,
        answer=payload.answer,
        priority=payload.priority,
    )
    return ORJSONResponse(item.model_dump(by_alias=True))


@router.patch("/qa/{pair_id}", response_class=ORJSONResponse)
async def knowledge_qa_update(
    request: Request,
    pair_id: str,
    payload: KnowledgeQAPayload,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Update a QA pair."""
    existing = await request.state.mongo.get_qa_pair(pair_id)
    if not existing:
        raise HTTPException(status_code=404, detail="QA pair not found")
        
    resolve_admin_project(request, existing.project)
    
    item = await request.state.mongo.update_qa_pair(
        pair_id,
        question=payload.question,
        answer=payload.answer,
        priority=payload.priority,
    )
    return ORJSONResponse(item.model_dump(by_alias=True))


@router.delete("/qa/{pair_id}", response_class=ORJSONResponse)
async def knowledge_qa_delete(
    request: Request,
    pair_id: str,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Delete a QA pair."""
    existing = await request.state.mongo.get_qa_pair(pair_id)
    if not existing:
        raise HTTPException(status_code=404, detail="QA pair not found")
        
    resolve_admin_project(request, existing.project)
    
    await request.state.mongo.delete_qa_pair(pair_id)
    return ORJSONResponse({"status": "deleted"})


@router.post("/qa/reorder", response_class=ORJSONResponse)
async def knowledge_qa_reorder(
    request: Request,
    payload: KnowledgeQAReorderPayload,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Reorder QA pairs."""
    # Verify access to all items
    for pair_id in payload.order:
        existing = await request.state.mongo.get_qa_pair(pair_id)
        if existing:
            resolve_admin_project(request, existing.project)
            
    await request.state.mongo.reorder_qa_pairs(payload.order)
    return ORJSONResponse({"status": "ok"})


@router.post("/qa/upload", response_class=ORJSONResponse, status_code=201)
async def knowledge_qa_upload(
    request: Request,
    project: str = Form(...),
    file: UploadFile = File(...),
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Import QA pairs from a file."""
    project_scope = resolve_admin_project(request, project, required=True)
    
    pairs = await _read_qa_upload(file)
    
    count = 0
    for pair in pairs:
        await request.state.mongo.create_qa_pair(
            project=project_scope,
            question=pair["question"],
            answer=pair["answer"],
            priority=pair.get("priority", 0),
        )
        count += 1
        
    return ORJSONResponse({"imported": count, "project": project_scope})


@router.get("/documents/{file_id}")
async def knowledge_download(
    request: Request,
    file_id: str,
    _: Any = Depends(require_admin),
) -> Response:
    """Return the raw contents of a document from GridFS."""
    
    mongo_cfg = MongoSettings()
    collection = getattr(request.state, "documents_collection", mongo_cfg.documents)
    
    try:
        meta, payload = await request.state.mongo.get_document_with_content(collection, file_id)
    except Exception as exc:
        logger.error("knowledge_download_failed", file_id=file_id, error=str(exc))
        raise HTTPException(status_code=404, detail="Document not found") from exc
        
    # Check access
    resolve_admin_project(request, meta.get("project"))
    
    filename = meta.get("name") or "document"
    content_type = meta.get("content_type") or "application/octet-stream"
    
    return Response(
        content=payload,
        media_type=content_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


# Unanswered Questions

@router.get("/unanswered", response_class=ORJSONResponse)
async def knowledge_unanswered_list(
    request: Request,
    project: str | None = None,
    limit: int = 100,
    offset: int = 0,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """List unanswered questions."""
    project_scope = resolve_admin_project(request, project)
    items = await request.state.mongo.list_unanswered(
        project=project_scope,
        limit=limit,
        offset=offset,
    )
    return ORJSONResponse({"items": [item.model_dump(by_alias=True) for item in items]})


@router.post("/unanswered/clear", response_class=ORJSONResponse)
async def knowledge_unanswered_clear(
    request: Request,
    payload: KnowledgeUnansweredClearPayload,
    _: Any = Depends(require_admin),
) -> ORJSONResponse:
    """Clear unanswered questions."""
    project_scope = resolve_admin_project(request, payload.project)
    count = await request.state.mongo.clear_unanswered(project_scope)
    return ORJSONResponse({"cleared": count})


# Service Management

@router.get("/service/status", response_class=ORJSONResponse)
async def knowledge_service_status(
    request: Request,
    _: Any = Depends(require_super_admin),
) -> ORJSONResponse:
    """Get knowledge service configuration and status."""
    doc = await request.state.mongo.get_setting(KNOWLEDGE_SERVICE_KEY) or {}
    
    # Defaults
    enabled = bool(doc.get("enabled", False))
    idle = int(doc.get("idle_threshold_seconds") or 300)
    poll = int(doc.get("poll_interval_seconds") or 60)
    cooldown = int(doc.get("cooldown_seconds") or 900)
    raw_mode = str(doc.get("mode") or "").strip().lower()
    mode = raw_mode if raw_mode in KNOWLEDGE_SERVICE_ALLOWED_MODES else KNOWLEDGE_SERVICE_DEFAULT_MODE
    prompt = doc.get("processing_prompt") or KNOWLEDGE_SERVICE_DEFAULT_PROMPT
    manual_msg = doc.get("manual_mode_message") or KNOWLEDGE_SERVICE_MANUAL_MESSAGE

    return ORJSONResponse({
        "enabled": enabled,
        "mode": mode,
        "processingPrompt": prompt,
        "idleThresholdSeconds": idle,
        "pollIntervalSeconds": poll,
        "cooldownSeconds": cooldown,
        "manualModeMessage": manual_msg,
    })


@router.post("/service/update", response_class=ORJSONResponse)
async def knowledge_service_update(
    request: Request,
    payload: KnowledgeServiceConfig,
    _: Any = Depends(require_super_admin),
) -> ORJSONResponse:
    """Update knowledge service configuration."""
    update_data = payload.model_dump(exclude_unset=True, by_alias=False)
    if not update_data:
        return await knowledge_service_status(request)

    await request.state.mongo.upsert_setting(KNOWLEDGE_SERVICE_KEY, update_data)
    return await knowledge_service_status(request)


@router.post("/service/run", response_class=ORJSONResponse)
async def knowledge_service_run(
    request: Request,
    payload: KnowledgeServiceRunRequest | None = None,
    _: Any = Depends(require_super_admin),
) -> ORJSONResponse:
    """Manually trigger knowledge service processing."""
    mode = payload.mode if payload else None
    force = payload.force if payload else False
    
    runner = KnowledgeServiceRunner(request.state.mongo)
    result = await runner.run_once(mode_override=mode, force=force)
    
    return ORJSONResponse(result)


@router.get("/processing/state", response_class=ORJSONResponse)
async def intelligent_processing_state(
    request: Request,
    _: Any = Depends(require_super_admin),
) -> ORJSONResponse:
    """Get intelligent processing state (prompt)."""
    doc = await request.state.mongo.get_setting(KNOWLEDGE_SERVICE_KEY) or {}
    prompt = doc.get("processing_prompt") or KNOWLEDGE_SERVICE_DEFAULT_PROMPT
    return ORJSONResponse({"prompt": prompt})


@router.post("/processing/prompt", response_class=ORJSONResponse)
async def intelligent_processing_save_prompt(
    request: Request,
    payload: IntelligentProcessingPromptPayload,
    _: Any = Depends(require_super_admin),
) -> ORJSONResponse:
    """Update intelligent processing prompt."""
    await request.state.mongo.upsert_setting(
        KNOWLEDGE_SERVICE_KEY,
        {"processing_prompt": payload.prompt}
    )
    return ORJSONResponse({"status": "ok", "prompt": payload.prompt})
